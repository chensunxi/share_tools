import imaplib
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import time
import datetime
import threading
import schedule
import email
import urllib3
import email.header
from email.header import decode_header
import traceback
from email.mime.base import MIMEBase
from email import encoders
from cryptography.fernet import Fernet
import os
from bs4 import BeautifulSoup
import zipfile
import sys
import ssl

from werkzeug.debug.repr import missing

from mail_monitor.download_archives import DownloadArchives
from utils.common_utils import extract_chinese,get_script_directory
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
from openpyxl.utils import get_column_letter

# 禁用证书警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 添加密钥常量
ENCRYPTION_KEY = b'YMS9XCt0Mk5ldHFWNWFSdjh3THNkTzZwZzR4RmF5RWE='  # 这是一个随机生成的base64编码密钥

class PasswordManager:
    def __init__(self):
        self.key = ENCRYPTION_KEY
        self.cipher_suite = Fernet(self.key)

    def encrypt(self, password):
        """加密密码"""
        return self.cipher_suite.encrypt(password.encode()).decode()

    def decrypt(self, encrypted_password):
        """解密密码"""
        try:
            return self.cipher_suite.decrypt(encrypted_password.encode()).decode()
        except:
            return encrypted_password  # 如果解密失败，返回原始值

class AutoReplyMail:
    # 文档类型映射表
    DOC_TYPE_MAP = {
        '身份证': 0,
        '学历证书': 1,
        '学位证书': 2,
        '职称证书': 3,
        '资格证书': 4,
        '获奖证书': 5,
        '工作经历': 6,
        '其他': 7,
        '照片': 8
    }

    def __init__(self, recv_monitor_mail, send_copy_mail, auto_reply_text='', auth_enter_mail=None, keywords=''):
        self.recv_monitor_mail = recv_monitor_mail
        self.send_copy_mail = send_copy_mail
        self.auto_reply_text = auto_reply_text
        self.auth_enter_mail = auth_enter_mail or []
        self.keywords = keywords
        self.is_running = False
        self.imap_connections = {}
        self.smtp_connections = {}
        self.password_manager = PasswordManager()
        self.monitor_threads = []
        
        # 延迟初始化连接
        self._init_success = False
        self.log("初始化完成，等待启动监控...")

    def log(self, message):
        """日志输出"""
        print(message)

    def init_connections(self):
        """初始化所有邮箱连接"""
        for config in self.recv_monitor_mail:
            email = config['email']
            password = self.password_manager.decrypt(config['password'])
            try:
                # IMAP连接
                imap = imaplib.IMAP4_SSL("imap.exmail.qq.com")
                imap.login(email, password)
                self.imap_connections[email] = imap
                
                # SMTP连接
                smtp = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
                smtp.login(email, password)
                self.smtp_connections[email] = smtp
                
                self.log(f"邮箱 {email} 连接成功")
            except Exception as e:
                self.log(f"邮箱 {email} 连接失败: {str(e)}")

    def reconnect_imap(self, email_config):
        """重新连接IMAP服务器"""
        try:
            email_addr = email_config['email']
            password = self.password_manager.decrypt(email_config['password'])
            
            # 先尝试关闭旧连接
            if email_addr in self.imap_connections:
                try:
                    self.imap_connections[email_addr].close()
                    self.imap_connections[email_addr].logout()
                except:
                    pass
                del self.imap_connections[email_addr]
            
            # 创建新连接
            imap = imaplib.IMAP4_SSL("imap.exmail.qq.com")
            imap.login(email_addr, password)
            self.imap_connections[email_addr] = imap
            self.log(f"成功重新连接到邮箱 {email_addr}")
            return imap
        except Exception as e:
            self.log(f"重新连接邮箱 {email_addr} 失败: {str(e)}")
            return None

    def check_connection(self, imap):
        """检查IMAP连接是否有效"""
        try:
            imap.noop()
            return True
        except:
            return False

    def check_and_reply(self, email_config):
        """检查邮件并回复"""
        try:
            email_addr = email_config['email']
            
            # 初始化连接
            if email_addr not in self.imap_connections:
                imap = self.reconnect_imap(email_config)
                if not imap:
                    self.log(f"无法连接到邮箱 {email_addr}，等待下次重试")
                    time.sleep(300)  # 等待5分钟后重试
                    return
            
            imap = self.imap_connections[email_addr]
            
            while self.is_running:
                try:
                    # 检查连接是否有效
                    if not self.check_connection(imap):
                        self.log(f"检测到连接已断开，尝试重新连接到邮箱 {email_addr}")
                        imap = self.reconnect_imap(email_config)
                        if not imap:
                            time.sleep(300)  # 等待5分钟后重试
                            continue
                    
                    # 每次检查前重新选择收件箱
                    try:
                        imap.select('INBOX')
                    except Exception as e:
                        if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                            self.log(f"选择收件箱时出现EOF错误，尝试重新连接")
                            imap = self.reconnect_imap(email_config)
                            if not imap:
                                time.sleep(300)
                                continue
                            imap.select('INBOX')
                    
                    # 搜索未读且未回复的邮件
                    try:
                        _, messages = imap.search(None, '(UNSEEN UNANSWERED)')
                    except Exception as e:
                        if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                            self.log(f"搜索邮件时出现EOF错误，尝试重新连接")
                            imap = self.reconnect_imap(email_config)
                            if not imap:
                                time.sleep(300)
                                continue
                            imap.select('INBOX')
                            _, messages = imap.search(None, '(UNSEEN UNANSWERED)')
                    
                    for num in messages[0].split():
                        if not self.is_running:
                            break
                            
                        try:
                            # 获取邮件内容
                            try:
                                _, msg_data = imap.fetch(num, '(RFC822)')
                            except Exception as e:
                                if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                                    self.log(f"获取邮件内容时出现EOF错误，尝试重新连接")
                                    imap = self.reconnect_imap(email_config)
                                    if not imap:
                                        time.sleep(300)
                                        continue
                                    imap.select('INBOX')
                                    _, msg_data = imap.fetch(num, '(RFC822)')
                                    
                            email_body = msg_data[0][1]
                            message = email.message_from_bytes(email_body)
                            
                            # 立即标记为已读，避免重复处理
                            try:
                                imap.store(num, '+FLAGS', r'(\Seen)')
                            except Exception as e:
                                if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                                    self.log(f"标记邮件时出现EOF错误，尝试重新连接")
                                    imap = self.reconnect_imap(email_config)
                                    if not imap:
                                        time.sleep(300)
                                        continue
                                    imap.select('INBOX')
                                    imap.store(num, '+FLAGS', r'(\Seen)')
                            
                            # 检查是否已回复
                            message_id = message.get('Message-ID')
                            if not message_id:
                                self.log("邮件没有 Message-ID，跳过处理")
                                continue
                            
                            if self.is_replied(message_id):
                                self.log("邮件已回复过，跳过处理")
                                continue
                            
                            # 处理邮件
                            self.process_email(message, email_config)
                            
                            # 确保标记已更新
                            try:
                                imap.select('INBOX')  # 重新选择收件箱以刷新状态
                            except Exception as e:
                                if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                                    self.log(f"刷新收件箱时出现EOF错误，尝试重新连接")
                                    imap = self.reconnect_imap(email_config)
                                    if not imap:
                                        time.sleep(300)
                                        continue
                            
                        except Exception as e:
                            self.log(f"处理邮件时出错: {str(e)}")
                            if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                                self.log("尝试重新连接...")
                                imap = self.reconnect_imap(email_config)
                                if not imap:
                                    time.sleep(300)
                                    continue
                            continue
                            
                    time.sleep(30)  # 每30秒检查一次
                    
                except Exception as e:
                    self.log(f"检查邮件时出错: {str(e)}")
                    if isinstance(e, (EOFError, ssl.SSLEOFError)) or 'EOF' in str(e):
                        self.log("尝试重新连接...")
                        imap = self.reconnect_imap(email_config)
                        if not imap:
                            time.sleep(300)
                            continue
                    else:
                        time.sleep(300)  # 其他错误等待5分钟再试
                    
        except Exception as e:
            self.log(f"邮件监控线程出错: {str(e)}")
        finally:
            try:
                if email_addr in self.imap_connections:
                    self.imap_connections[email_addr].close()
                    self.imap_connections[email_addr].logout()
                    del self.imap_connections[email_addr]
            except:
                pass

    def is_replied(self, message_id):
        """检查邮件是否已回复"""
        try:
            for email, imap in self.imap_connections.items():
                try:
                    imap.select('INBOX')
                    # 使用 Message-ID 和 \Answered 标记来检查
                    _, messages = imap.search(None, f'(HEADER "Message-ID" "{message_id}" ANSWERED)')
                    if messages[0]:
                        self.log(f"邮件已经回复过，Message-ID: {message_id}")
                        # 验证标记
                        for num in messages[0].split():
                            _, flags = imap.fetch(num, '(FLAGS)')
                            self.log(f"邮件标记状态: {flags[0].decode()}")
                        return True
                except Exception as e:
                    self.log(f"在邮箱 {email} 中检查邮件状态时出错: {str(e)}")
            return False
        except Exception as e:
            self.log(f"检查邮件回复状态时出错: {str(e)}")
            return False

    def mark_as_replied(self, message_id):
        """标记邮件为已回复"""
        try:
            for email, imap in self.imap_connections.items():
                try:
                    imap.select('INBOX')
                    # 查找对应的邮件
                    _, messages = imap.search(None, f'HEADER "Message-ID" "{message_id}"')
                    if messages[0]:  # 确保找到了邮件
                        for num in messages[0].split():
                            # 标记为已回复和已读
                            imap.store(num, '+FLAGS', r'(\Seen \Answered)')
                            self.log(f"已将邮件标记为已回复和已读，Message-ID: {message_id}")
                            # 验证标记是否成功
                            _, flags = imap.fetch(num, '(FLAGS)')
                            self.log(f"邮件当前标记: {flags[0].decode()}")
                            return True
                except Exception as e:
                    self.log(f"在邮箱 {email} 中标记邮件状态时出错: {str(e)}")
            return False
        except Exception as e:
            self.log(f"标记邮件状态时出错: {str(e)}")
            return False

    def need_reply(self, message):
        """检查是否需要回复邮件"""
        try:
            # 获取发件人邮箱
            from_addr = self.get_email_address(message['From'])
            
            # 检查发件人是否在准入邮箱列表中
            authorized_emails = [auth['email'] for auth in self.auth_enter_mail]
            if from_addr not in authorized_emails:
                self.log(f"发件人 {from_addr} 不在准入邮箱列表中，跳过回复")
                return False
            
            # 获取并解码邮件主题
            subject = self.decode_header_str(message['Subject'])
            self.log(f"原始邮件主题: {message['Subject']}")
            self.log(f"解码后的主题: {subject}")
            
            # 检查主题中是否包含关键词
            for keyword in self.keywords.split(','):
                keyword = keyword.strip()
                if keyword and keyword in subject:
                    self.log(f"找到关键词: {keyword}")
                    return True
            
            return False
        except Exception as e:
            self.log(f"检查邮件是否需要回复时出错: {str(e)}")
            return False

    def send_reply(self, original_message, email_config=None):
        """发送回复邮件"""
        try:
            # 获取原始邮件信息
            subject = self.decode_header_str(original_message['Subject'])
            from_addr = self.get_email_address(original_message['From'])
            
            # 使用接收邮件的邮箱配置
            if not email_config:
                self.log("没有提供邮箱配置，无法回复邮件")
                return False
            
            reply_email = email_config['email']
            reply_password = self.password_manager.decrypt(email_config['password'])
            
            # 创建回复邮件
            reply = MIMEMultipart()
            reply['From'] = reply_email
            reply['To'] = from_addr
            reply['Subject'] = f"auto reply: {subject}"
            
            # 获取邮件内容并解析表格数据
            content = self.get_email_content(original_message)
            table_data = self.parse_table_data(content) if content else []
            
            # 构建回复内容
            html_content = self.prepare_reply_email(original_message, table_data)
            
            # 添加HTML内容
            reply.attach(MIMEText(html_content, 'html', 'utf-8'))
            
            # 发送邮件
            try:
                smtp = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
                smtp.login(reply_email, reply_password)
                
                recipients = [from_addr]
                
                smtp.sendmail(reply_email, recipients, reply.as_string())
                smtp.quit()
                
                self.log(f"已发送回复邮件给 {from_addr}")
                return True
                
            except Exception as e:
                self.log(f"发送回复邮件时出错: {str(e)}")
                return False
            
        except Exception as e:
            self.log(f"准备回复邮件时出错: {str(e)}")
            return False

    def decode_header_str(self, header):
        """解码邮件头信息"""
        try:
            # 如果是空值，返回空字符串
            if not header:
                return ''
            
            # 解码邮件主题
            decoded_header = decode_header(header)
            header_parts = []
            
            for content, charset in decoded_header:
                if isinstance(content, bytes):
                    try:
                        # 尝试使用指定的字符集
                        if charset:
                            header_parts.append(content.decode(charset))
                        else:
                            # 如果没有指定字符集，尝试常用字符集
                            for encoding in ['utf-8', 'gb18030', 'gbk', 'gb2312']:
                                try:
                                    header_parts.append(content.decode(encoding))
                                    break
                                except:
                                    continue
                    except:
                        # 如果解码失败，使用原始内容
                        header_parts.append(str(content))
                else:
                    header_parts.append(str(content))
                
            return ' '.join(header_parts)
        except Exception as e:
            self.log(f"解码邮件头信息时出错: {str(e)}")
            return header

    def get_email_address(self, addr_str):
        """从邮件地址字符串中提取邮箱地址"""
        try:
            # 处理 "Name" <email@domain.com> 格式
            if '<' in addr_str and '>' in addr_str:
                start = addr_str.find('<') + 1
                end = addr_str.find('>')
                return addr_str[start:end]
            return addr_str.strip()
        except:
            return addr_str

    def start_monitoring(self, mode='realtime', schedule_time='09:00'):
        """启动邮件监控"""
        try:
            self.stop_monitoring()  # 先停止现有监控
            self.is_running = True
            self.monitor_threads = []  # 重置线程列表
            
            if mode == 'realtime':
                self.log("开始实时监控...")
                for email_config in self.recv_monitor_mail:
                    email = email_config['email']
                    password = self.password_manager.decrypt(email_config['password'])
                    
                    # 测试邮箱连接
                    try:
                        imap = imaplib.IMAP4_SSL("imap.exmail.qq.com")
                        imap.login(email, password)
                        imap.select('INBOX')
                        self.log(f"邮箱 {email} 连接成功")
                        imap.logout()
                    except Exception as e:
                        self.log(f"邮箱 {email} 连接失败: {str(e)}")
                        continue
                    
                    # 创建监控线程
                    thread = threading.Thread(
                        target=self.check_and_reply,
                        args=(email_config,),  # 传递 email_config 参数
                        daemon=True
                    )
                    thread.start()
                    self.monitor_threads.append(thread)
                
                return True
            elif mode == 'scheduled':
                self.log(f"启动定时监控模式，执行时间: {schedule_time}")
                schedule.every().day.at(schedule_time).do(self.check_and_reply)
                self._scheduled_monitoring()
        except Exception as e:
            self.log(f"启动监控失败: {str(e)}")
            self.log(f"错误详情: {traceback.format_exc()}")
            return False

    def _realtime_monitoring(self):
        """实时监控处理"""
        while self.is_running:
            try:
                self.check_and_reply()
                time.sleep(60)  # 每分钟检查一次
            except Exception as e:
                self.log(f"监控过程出错: {str(e)}")
                time.sleep(300)  # 出错后等待5分钟再试

    def _scheduled_monitoring(self):
        """定时监控处理"""
        while self.is_running:
            schedule.run_pending()
            time.sleep(60)

    def stop_monitoring(self):
        """停止监控"""
        self.is_running = False
        # 等待所有监控线程结束
        for thread in self.monitor_threads:
            if thread.is_alive():
                thread.join(timeout=5)
        self.monitor_threads = []
        
        # 关闭所有连接
        for email in self.imap_connections:
            try:
                self.imap_connections[email].close()
                self.imap_connections[email].logout()
            except:
                pass
        for email in self.smtp_connections:
            try:
                self.smtp_connections[email].quit()
            except:
                pass
        self.log("邮件监控已停止")

    def update_config(self, new_config):
        """更新配置"""
        try:
            # 更新配置
            self.recv_monitor_mail = new_config.get('recv_monitor_mail', [])
            self.send_copy_mail = new_config.get('send_copy_mail', [])
            self.keywords = new_config.get('keywords', '')  # 确保更新关键词
            self.auto_reply_text = new_config.get('auto_reply_text', '')
            self.auth_enter_mail = new_config.get('auth_enter_mail', [])
            
            self.log("配置已更新")
        except Exception as e:
            self.log(f"更新配置失败: {str(e)}")

    def get_email_content(self, message):
        """获取邮件内容"""
        try:
            content = ""
            if message.is_multipart():
                for part in message.walk():
                    if part.get_content_type() in ["text/html", "text/plain"]:
                        # 获取编码
                        charset = part.get_content_charset()
                        if not charset:
                            charset = 'gb18030'  # 如果没有指定编码，默认使用gb18030
                        
                        # 获取内容并解码
                        try:
                            payload = part.get_payload(decode=True)
                            # 尝试不同的编码
                            for encoding in [charset, 'gb18030', 'gbk', 'utf-8', 'gb2312']:
                                try:
                                    content = payload.decode(encoding)
                                    self.log(f"成功使用 {encoding} 解码邮件内容")
                                    if part.get_content_type() == "text/html":
                                        return content
                                    break
                                except UnicodeDecodeError:
                                    continue
                        except Exception as e:
                            self.log(f"解码邮件部分内容时出错: {str(e)}")
                            continue
            else:
                # 获取编码
                charset = message.get_content_charset()
                if not charset:
                    charset = 'gb18030'  # 如果没有指定编码，默认使用gb18030
                
                # 获取内容并解码
                try:
                    payload = message.get_payload(decode=True)
                    # 尝试不同的编码
                    for encoding in [charset, 'gb18030', 'gbk', 'utf-8', 'gb2312']:
                        try:
                            content = payload.decode(encoding)
                            self.log(f"成功使用 {encoding} 解码邮件内容")
                            break
                        except UnicodeDecodeError:
                            continue
                except Exception as e:
                    self.log(f"解码邮件内容时出错: {str(e)}")
            
            if not content:
                self.log("未能成功解码邮件内容")
            return content
            
        except Exception as e:
            self.log(f"获取邮件内容时出错: {str(e)}")
            return ""

    def parse_table_data(self, html_content):
        """解析邮件中的表格数据"""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            tables = soup.find_all('table')
            
            results = []
            for table in tables:
                rows = table.find_all('tr')
                headers = None
                
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if not cells:
                        continue
                    
                    # 获取表头
                    if not headers:
                        headers = []
                        for cell in cells:
                            header = cell.get_text(strip=True)
                            headers.append(header)
                        continue
                    
                    # 查找员工编号、管理主体、所需资料的列索引
                    emp_id_idx = next((i for i, h in enumerate(headers) if '员工编号' in h), None)
                    emp_name_idx = next((i for i, h in enumerate(headers) if '工作名' in h), None)
                    org_idx = next((i for i, h in enumerate(headers) if '管理主体' in h), None)
                    emp_dept1 = next((i for i, h in enumerate(headers) if '管理关系一级部门' in h), None)
                    emp_dept2 = next((i for i, h in enumerate(headers) if '管理关系二级部门' in h), None)
                    materials_idx = next((i for i, h in enumerate(headers) if '所需资料' in h), None)
                    
                    if all(idx is not None for idx in [emp_id_idx, org_idx, materials_idx]):
                        row_data = [cell.get_text(strip=True) for cell in cells]
                        if len(row_data) > max(emp_id_idx, emp_name_idx, org_idx, materials_idx):
                            result = {
                                'emp_id': row_data[emp_id_idx],
                                'emp_name': row_data[emp_name_idx],
                                'emp_org': row_data[org_idx],
                                'emp_dept1': row_data[emp_dept1],
                                'emp_dept2': row_data[emp_dept2],
                                'emp_archives': row_data[materials_idx]
                            }
                            results.append(result)
            
            # 打印解析结果
            if results:
                self.log("\n=== 邮件表格解析结果 ===")
                for idx, item in enumerate(results, 1):
                    self.log(f"\n条目 {idx}:")
                    self.log(f"员工编号: {item['emp_id']}")
                    self.log(f"管理主体: {item['emp_org']}")
                    self.log(f"所需资料: {item['emp_archives']}")
                self.log("=====================")
            else:
                self.log("未在邮件中找到有效的表格数据")
            
            return results
            
        except Exception as e:
            self.log(f"解析表格数据时出错: {str(e)}")
            return []

    def check_download_success(self, user_data, download_scope='000000000'):
        """检查员工文档下载是否完全成功
        
        Args:
            user_data (dict): 员工数据
            download_scope (str): 下载范围标记字符串
        
        Returns:
            tuple: (是否全部成功, 缺失的文档列表)
        """
        required_docs = user_data.get('emp_archives', '').split('、')
        missing_docs = []
        
        for doc in required_docs:
            if doc and doc in self.DOC_TYPE_MAP:
                idx = self.DOC_TYPE_MAP[doc]
                if download_scope[idx] != '1':
                    missing_docs.append(doc)
        
        return len(missing_docs) == 0, missing_docs

    def record_download_result(self, user_data, download_dir, success, error_msg, reply_sent, download_scope='000000000'):
        """记录下载结果到Excel文件"""
        try:
            # 获取当前年月
            current_date = datetime.datetime.now()
            year_month = current_date.strftime('%Y%m')

            # 构建Excel文件路径（使用绝对路径）
            excel_file = os.path.join(os.path.dirname(download_dir), f'mail_report_{year_month}.xlsx')

            # 准备记录数据
            record = {
                '序号': None,  # 将在后面设置
                '日期': current_date.strftime('%Y%m%d'),
                '员工编号': user_data.get('员工编号', ''),
                '工作名': user_data.get('工作名', ''),
                '管理主体': user_data.get('管理主体', ''),
                '发件邮箱': user_data.get('sender_email', ''),
                '邮件标题': user_data.get('subject', ''),
                '是否回复': '√' if reply_sent else '×',  # 使用√和×替代是和否
                '所需资料': user_data.get('所需资料', ''),
                '是否成功': '√' if success else '×',  # 使用√和×替代是和否
                '错误信息': error_msg
            }

            # 如果文件存在，读取现有数据
            if os.path.exists(excel_file):
                df = pd.read_excel(excel_file)
            else:
                df = pd.DataFrame(columns=['序号', '日期', '员工编号', '工作名', '管理主体',
                                         '发件邮箱', '邮件标题', '是否回复', '所需资料', '是否成功', '错误信息'])

            # 添加新记录
            new_record = pd.DataFrame([record])
            df = pd.concat([df, new_record], ignore_index=True)

            # 更新序号
            df['序号'] = range(1, len(df) + 1)

            # 确保员工编号格式正确（5位数，左侧补0）
            df['员工编号'] = df['员工编号'].astype(str).str.zfill(5)

            # 保存到Excel文件
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

                # 获取工作表
                worksheet = writer.sheets['Sheet1']

                # 设置列宽
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    ) + 2  # 添加一些额外空间
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length

                # 设置样式
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(name="微软雅黑", size=8, color="FFFFFF", bold=True)
                data_font = Font(name="微软雅黑", size=8)

                # 设置边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 应用标题行样式
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 应用数据行样式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = data_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 设置"是否成功"和"是否回复"列的条件格式
                success_font = Font(name="微软雅黑", size=8, color="00B050")  # 绿色
                failure_font = Font(name="微软雅黑", size=8, color="FF0000")  # 红色

                for row in range(2, worksheet.max_row + 1):
                    # 处理"是否成功"列
                    success_cell = worksheet.cell(row=row, column=df.columns.get_loc('是否成功') + 1)
                    if success_cell.value == '√':
                        success_cell.font = success_font
                    else:
                        success_cell.font = failure_font

                    # 处理"是否回复"列
                    reply_cell = worksheet.cell(row=row, column=df.columns.get_loc('是否回复') + 1)
                    if reply_cell.value == '√':
                        reply_cell.font = success_font
                    else:
                        reply_cell.font = failure_font

            self.log(f"已记录下载结果到: {excel_file}")


            # 检查下载结果
            all_success, missing_docs = self.check_download_success(user_data, download_scope)
            
            # 如果有缺失文档,添加到通知列表
            if missing_docs:
                self.missing_docs_notifications.append({
                    'business': user_data.get('emp_org', ''),
                    'name': user_data.get('emp_name', ''),
                    'email': user_data.get('sender_email', ''),
                    'missing_docs': '、'.join(missing_docs)
                })
                
        except Exception as e:
            self.log(f"记录下载结果时出错: {str(e)}")
            traceback.print_exc()

    def prepare_reply_email(self, original_message, table_data=None, download_dir=None, download_results=None):
        """准备回复邮件内容"""
        try:
            # 获取原始邮件信息
            subject = self.decode_header_str(original_message['Subject'])
            from_addr = self.get_email_address(original_message['From'])
            
            # 获取邮件内容并解析表格数据
            content = self.get_email_content(original_message)
            if not table_data and content:
                table_data = self.parse_table_data(content)
            
            # 创建回复内容
            html_content = self.auto_reply_text
            
            # 如果有表格数据,添加处理结果表格
            if table_data:
                result_table = self.create_reply_table(table_data, download_dir, download_results)
                if result_table:
                    html_content += "<br><br>" + result_table
            
            # 替换标签
            if "{$manager}" in html_content and download_results:
                # 收集需要通知的业务单元
                business_units = set()
                for result in download_results:
                    # 检查是否有缺失文档
                    all_success, _ = self.check_download_success(result, result.get('download_scope', '000000000'))
                    if not all_success:
                        business = result.get('emp_org', '')
                        if business:
                            business_units.add(business)
                
                # 从抄送列表中获取对应的管理者姓名
                manager_names = []
                for cc in self.send_copy_mail:
                    if cc['business'] in business_units:
                        manager_names.append(f"{cc['business']}-{cc['name']}")
                
                # 替换标签
                html_content = html_content.replace("{$manager}", "、".join(manager_names))
            
            return html_content
            
        except Exception as e:
            self.log(f"准备回复邮件内容时出错: {str(e)}")
            traceback.print_exc()
            return None

    def create_reply_table(self, table_data, download_dir=None, download_results=None):
        """创建回复表格"""
        try:
            # 表头
            html = """
            <table border="1" cellspacing="0" cellpadding="5" style="border-collapse: collapse;">
                <tr style="background-color: #f2f2f2;">
                    <th>序号</th>
                    <th>姓名</th>
                    <th>所属单位</th>
                    <th>所需资料</th>
                    <th>处理结果</th>
                </tr>
            """
            
            # 遍历表格数据
            for i, row in enumerate(table_data, 1):
                # 查找对应的下载结果
                download_result = None
                if download_results:
                    for result in download_results:
                        if (result.get('name') == row.get('emp_name') and 
                            result.get('business') == row.get('emp_org')):
                            download_result = result
                            break
                
                # 检查下载结果
                if download_result:
                    all_success, missing_docs = self.check_download_success(
                        row, 
                        download_result.get('download_scope', '000000000')
                    )
                    status = "已完成" if all_success else f"缺少: {', '.join(missing_docs)}"
                else:
                    status = "处理失败"
                
                # 添加行
                html += f"""
                <tr>
                    <td align="center">{i}</td>
                    <td>{row.get('emp_name', '')}</td>
                    <td>{row.get('emp_org', '')}</td>
                    <td>{row.get('emp_archives', '')}</td>
                    <td>{status}</td>
                </tr>
                """
            
            html += "</table>"
            return html
            
        except Exception as e:
            self.log(f"创建回复表格时出错: {str(e)}")
            traceback.print_exc()
            return None

    def process_email(self, message, email_config):
        """处理单个邮件"""
        try:
            # 检查是否已回复
            message_id = message['Message-ID']
            if self.is_replied(message_id):
                self.log("邮件已回复，跳过处理")
                return

            # 获取邮件内容
            email_content = self.get_email_content(message)
            if not email_content:
                self.log("未找到邮件内容，跳过处理")
                return

            # 解析表格数据
            table_data = self.parse_table_data(email_content)
            
            if not table_data:
                self.log("未找到有效的表格数据，跳过处理")
                return
            
            # 创建下载目录（使用绝对路径）
            current_time = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            # 使用项目根目录而不是脚本目录
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            download_dir = os.path.abspath(os.path.join(base_dir, 'download', 'mail_attach', current_time))
            os.makedirs(download_dir, exist_ok=True)
            self.log(f"创建下载目录：{download_dir}")
            
            # 下载文件
            download_success = False
            error_msg = ""
            reply_sent = False
            zip_file_path = None
            download_results = {}  # 存储每个员工的下载结果
            
            try:
                # 准备所有用户数据
                users_info = []
                for user_data in table_data:
                    # 解析所需资料并设置 download_scope
                    required_docs = user_data['emp_archives'].split('、')
                    download_scope = ['0'] * 9  # 初始化为9个0
                    
                    # 根据所需资料设置对应位置为1
                    for doc in required_docs:
                        for key, index in self.DOC_TYPE_MAP.items():
                            if key in doc:
                                download_scope[index] = '1'
                    
                    user_info = {
                        'job_no': user_data['emp_id'],
                        'job_name': user_data.get('emp_name'),
                        'business': user_data['emp_org'],
                        'required_docs': user_data['emp_archives'],
                        'sender_email': self.get_email_address(message['From']),
                        'subject': self.decode_header_str(message['Subject']),
                        'download_scope': ''.join(download_scope)
                    }
                    users_info.append(user_info)
                
                # 创建下载档案实例
                downloader = DownloadArchives(
                    email_config.get('erpuser', ''),
                    self.password_manager.decrypt(email_config.get('erppwd', '')),
                    download_dir,
                    users_info,
                    None  # 这里传入了 None 作为 out_logger
                )
                
                # 执行下载并获取下载结果
                download_results = downloader.run(0)  # 0表示测试环境
                
                # 检查是否有任何员工的资料下载成功
                has_files = False
                for root, dirs, files in os.walk(download_dir):
                    if files:
                        has_files = True
                        break
                
                # 准备回复邮件
                reply_prepared = self.prepare_reply_email(message, table_data, download_dir, download_results)
                
                if has_files:
                    # 创建压缩文件
                    zip_file_path = os.path.join(download_dir, f'员工档案_{current_time}.zip')
                    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                        for root, dirs, files in os.walk(download_dir):
                            for file in files:
                                if not file.endswith('.zip'):  # 不要将zip文件本身添加到压缩包中
                                    file_path = os.path.join(root, file)
                                    arcname = os.path.relpath(file_path, download_dir)
                                    zipf.write(file_path, arcname)
                
                # 发送回复邮件
                if zip_file_path and os.path.exists(zip_file_path):  # 只有在有压缩文件时才附加文件
                    if self.send_reply_with_attachment(message, zip_file_path, reply_prepared, email_config):
                        self.log("回复邮件发送成功")
                        reply_sent = True
                        self.mark_as_replied(message_id)
                    else:
                        self.log("发送回复邮件失败")
                        error_msg = "发送回复邮件失败"
                else:  # 如果没有文件，发送不带附件的回复
                    if self.send_reply(message, email_config):
                        self.log("回复邮件（无附件）发送成功")
                        reply_sent = True
                        self.mark_as_replied(message_id)
                    else:
                        self.log("发送回复邮件失败")
                        error_msg = "发送回复邮件失败"
                
                # 在确认邮件发送状态后记录结果
                for user_data in table_data:
                    # 添加发件人和主题信息
                    user_data['sender_email'] = self.get_email_address(message['From'])
                    user_data['subject'] = self.decode_header_str(message['Subject'])
                    
                    # 获取该员工的下载结果
                    job_no = user_data['emp_id']
                    user_download_scope = download_results.get(job_no, '000000000')
                    
                    # 判断该员工的下载是否完全成功
                    user_download_success = True
                    required_docs = user_data['emp_archives'].split('、')
                    for doc in required_docs:
                        doc_found = False
                        for key, index in self.DOC_TYPE_MAP.items():
                            if key in doc and user_download_scope[index] == '1':
                                doc_found = True
                                break
                        if not doc_found:
                            user_download_success = False
                            break
                    
                    self.record_download_result(
                        user_data=user_data,
                        download_dir=download_dir,
                        success=user_download_success,
                        error_msg=error_msg,
                        reply_sent=reply_sent,
                        download_scope=user_download_scope
                    )
                
                # 检查是否有缺失资料的员工，如果有则发送提醒邮件
                missing_data = []
                for user_data in table_data:
                    job_no = user_data['emp_id']
                    user_download_scope = download_results.get(job_no, '000000000')
                    
                    # 判断该员工的下载是否完全成功
                    required_docs = user_data['emp_archives'].split('、')
                    missing_docs = []
                    
                    for doc in required_docs:
                        doc_found = False
                        for key, index in self.DOC_TYPE_MAP.items():
                            if key in doc and user_download_scope[index] == '1':
                                doc_found = True
                                break
                        if not doc_found:
                            missing_docs.append(doc)
                    
                    if missing_docs:
                        missing_data.append({
                            'emp_id': user_data['emp_id'],
                            'emp_name': user_data.get('emp_name', '未知'),
                            'emp_org': user_data['emp_org'],
                            'emp_dept1': user_data.get('emp_dept1', ''),
                            'emp_dept2': user_data.get('emp_dept2', ''),
                            'missing_docs': '、'.join(missing_docs)
                        })
                
                if missing_data:
                    self.send_missing_docs_notification(missing_data, email_config)
                
            except Exception as e:
                self.log(f"处理邮件时出错: {str(e)}")
                error_msg = str(e)
                
                # 即使出错也要记录结果
                for user_data in table_data:
                    # 添加发件人和主题信息
                    user_data['sender_email'] = self.get_email_address(message['From'])
                    user_data['subject'] = self.decode_header_str(message['Subject'])
                    
                    # 获取该员工的下载结果（如果有）
                    job_no = user_data['emp_id']
                    user_download_scope = download_results.get(job_no, '000000000')
                    
                    self.record_download_result(
                        user_data=user_data,
                        download_dir=download_dir,
                        success=False,
                        error_msg=error_msg,
                        reply_sent=reply_sent,
                        download_scope=user_download_scope
                    )
                
                # 如果出错且还没有发送回复，尝试发送不带附件的回复
                if not reply_sent:
                    reply_prepared = self.prepare_reply_email(message, table_data, download_dir, download_results)
                    if self.send_reply(message, email_config):
                        self.log("回复邮件（无附件）发送成功")
                        reply_sent = True
                        self.mark_as_replied(message_id)
                        
                        # 更新记录中的回复状态
                        for user_data in table_data:
                            # 获取该员工的下载结果（如果有）
                            job_no = user_data['emp_id']
                            user_download_scope = download_results.get(job_no, '000000000')
                            
                            self.record_download_result(
                                user_data=user_data,
                                download_dir=download_dir,
                                success=False,
                                error_msg=error_msg,
                                reply_sent=True,
                                download_scope=user_download_scope
                            )
                
                # 检查是否有缺失资料的员工，如果有则发送提醒邮件
                missing_data = []
                for user_data in table_data:
                    job_no = user_data['emp_id']
                    user_download_scope = download_results.get(job_no, '000000000')
                    
                    # 判断该员工的下载是否完全成功
                    required_docs = user_data['emp_archives'].split('、')
                    missing_docs = []
                    
                    for doc in required_docs:
                        doc_found = False
                        for key, index in self.DOC_TYPE_MAP.items():
                            if key in doc and user_download_scope[index] == '1':
                                doc_found = True
                                break
                        if not doc_found:
                            missing_docs.append(doc)
                    
                    if missing_docs:
                        missing_data.append({
                            'emp_id': user_data['emp_id'],
                            'emp_name': user_data.get('emp_name', '未知'),
                            'emp_org': user_data['emp_org'],
                            'emp_dept1': user_data.get('emp_dept1', ''),
                            'emp_dept2': user_data.get('emp_dept2', ''),
                            'missing_docs': '、'.join(missing_docs)
                        })
                
                if missing_data:
                    self.send_missing_docs_notification(missing_data, email_config)
            
        except Exception as e:
            self.log(f"处理邮件时出错: {str(e)}")
            traceback.print_exc()

    def send_reply_with_attachment(self, original_message, attachment_path, reply_prepared, email_config=None):
        """发送带附件的回复邮件"""
        try:
            # 获取原始邮件信息
            subject = self.decode_header_str(original_message['Subject'])
            from_addr = self.get_email_address(original_message['From'])
            
            # 使用接收邮件的邮箱配置
            if not email_config:
                self.log("没有提供邮箱配置，无法回复邮件")
                return False
            
            reply_email = email_config['email']
            reply_password = self.password_manager.decrypt(email_config['password'])
            
            # 创建回复邮件
            reply = MIMEMultipart()
            reply['From'] = reply_email
            reply['To'] = from_addr
            reply['Subject'] = f"Re: {subject}"
            
            # 获取邮件内容并解析表格数据
            content = self.get_email_content(original_message)
            table_data = self.parse_table_data(content) if content else []
            
            # 构建回复内容
            html_content = reply_prepared
            
            # 添加HTML内容
            reply.attach(MIMEText(html_content, 'html', 'utf-8'))
            
            # 添加附件
            if os.path.exists(attachment_path):
                with open(attachment_path, 'rb') as f:
                    attachment = MIMEBase('application', 'zip')
                    attachment.set_payload(f.read())
                    encoders.encode_base64(attachment)
                    attachment_name = os.path.basename(attachment_path)
                    attachment.add_header('Content-Disposition', 'attachment', 
                                       filename=attachment_name)
                    reply.attach(attachment)
            
            # 发送邮件
            try:
                smtp = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
                smtp.login(reply_email, reply_password)
                
                recipients = [from_addr]
                
                smtp.sendmail(reply_email, recipients, reply.as_string())
                smtp.quit()
                
                self.log(f"已发送带附件的回复邮件给 {from_addr}")
                return True
                
            except Exception as e:
                self.log(f"发送回复邮件时出错: {str(e)}")
                return False
            
        except Exception as e:
            self.log(f"准备回复邮件时出错: {str(e)}")
            return False

    def send_missing_docs_notification(self, missing_data, email_config):
        """发送缺失资料提醒邮件"""
        try:
            if not missing_data or not email_config:
                return False
                
            # 获取邮箱配置
            reply_email = email_config['email']
            reply_password = self.password_manager.decrypt(email_config['password'])
            
            # 创建邮件
            msg = MIMEMultipart()
            msg['From'] = reply_email
            msg['Subject'] = "[提醒]如下员工投标\\入场\\审计\\项目检查资料缺失，请尽快提供！"
            
            # 收集所有缺失资料员工的管理主体
            business_units = set()
            for item in missing_data:
                business_units.add(item['emp_org'])
            
            # 根据管理主体筛选抄送邮箱
            recipients = []
            for business in business_units:
                for cc in self.send_copy_mail:
                    if cc['business'] == business:
                        recipients.append(cc['email'])
            
            if not recipients:
                self.log("没有找到匹配的抄送邮箱，无法发送提醒邮件")
                return False
                
            msg['To'] = ', '.join(recipients)
            
            # 构建邮件内容
            html_content = """
            <html>
            <body>
                <p>您好!</p>
                <p>如下员工入场\\审计\\项目检查相关资料暂未查询到，请尽快确认并提供给相关同事，谢谢！</p>
            """
            
            # 添加表格
            html_content += '<table border="1" cellspacing="0" cellpadding="5" style="border-collapse: collapse;">'
            # 添加表头
            html_content += '''
                <tr>
                    <th style="background-color: #4472C4; color: white;">员工编号</th>
                    <th style="background-color: #4472C4; color: white;">工作名</th>
                    <th style="background-color: #4472C4; color: white;">管理主体</th>
                    <th style="background-color: #4472C4; color: white;">管理关系一级部门</th>
                    <th style="background-color: #4472C4; color: white;">管理关系二级部门</th>
                    <th style="background-color: #4472C4; color: white;">缺失资料</th>
                </tr>
            '''
            
            # 添加数据行
            for item in missing_data:
                html_content += f'''
                    <tr>
                        <td>{item['emp_id']}</td>
                        <td>{item.get('emp_name', '未知')}</td>
                        <td>{item['emp_org']}</td>
                        <td>{item.get('emp_dept1', '')}</td>
                        <td>{item.get('emp_dept2', '')}</td>
                        <td>{item['missing_docs']}</td>
                    </tr>
                '''
            
            html_content += '''
                </table>
                </body>
                </html>
            '''
            
            # 添加HTML内容
            msg.attach(MIMEText(html_content, 'html', 'utf-8'))
            
            # 发送邮件
            try:
                smtp = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
                smtp.login(reply_email, reply_password)
                smtp.sendmail(reply_email, recipients, msg.as_string())
                smtp.quit()
                
                self.log(f"已发送缺失资料提醒邮件给 {', '.join(recipients)}")
                return True
                
            except Exception as e:
                self.log(f"发送缺失资料提醒邮件时出错: {str(e)}")
                return False
            
        except Exception as e:
            self.log(f"准备缺失资料提醒邮件时出错: {str(e)}")
            return False

def main():
    try:
        # 配置信息
        recv_monitor_mail = [{"email": "chensx@sunline.cn", "password": "ayjnmS3bckNZBthR"}]  # 可以添加多个邮箱
        send_copy_mail = ["chensunxi@163.com"]  # 抄送地址
        
        # 创建自动回复实例
        auto_reply = AutoReplyMail(
            recv_monitor_mail=recv_monitor_mail,
            send_copy_mail=send_copy_mail
        )
        
        # 选择监控模式：
        # 1. 实时监控
        auto_reply.start_monitoring(mode='realtime')
        # 2. 定时监控（比如每天9点执行）
        # auto_reply.start_monitoring(mode='scheduled', schedule_time="09:00")

    except KeyboardInterrupt:
        print("\n程序被用户中断")
    except Exception as e:
        print(f"程序出错: {str(e)}")
    finally:
        if 'auto_reply' in locals():
            auto_reply.stop_monitoring()

if __name__ == "__main__":
    main() 